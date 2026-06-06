# 사람 κ 검수 시트 생성 (★1 방어용)
#   라벨을 '가린' 채 댓글+직전 가격맥락만 보여주는 review_sheet.xlsx 생성.
#   정답(파이프라인 라벨)은 review_key.csv 에 따로 숨겨둠(검수자 편향 방지).
#   사람이 시트의 [사람_시점관계]·[사람_예측_방향]만 드롭다운으로 채우면 됨.
#   채운 뒤: python3 compute_kappa.py
import os
import glob
import argparse
import pandas as pd

from price_context import context_text, ROOT
from run_label import FILES  # train/test 4파일명

DATA = os.path.join(ROOT, "data")
N_DEFAULT = 250
SEED = 42


def load_pool() -> pd.DataFrame:
    """final_*.csv 우선(있으면 Class로 층화), 없으면 labeled_+source 병합."""
    finals = [os.path.join(DATA, f"final_{f}.csv") for f in FILES]
    if all(os.path.exists(p) for p in finals):
        df = pd.concat([pd.read_csv(p, encoding="utf-8-sig") for p in finals],
                       ignore_index=True)
        df["_strata"] = df["Class"]
        return df
    # fallback: 라벨링만 끝났을 때
    parts = []
    for f in FILES:
        lp = os.path.join(DATA, f"labeled_{f}.csv")
        if not os.path.exists(lp):
            continue
        lab = pd.read_csv(lp, encoding="utf-8-sig")
        src = pd.read_csv(os.path.join(DATA, f"{f}.csv"), encoding="utf-8-sig")
        m = src.merge(lab, on="commentId", how="inner")
        parts.append(m)
    df = pd.concat(parts, ignore_index=True)
    df = df[df["시점관계"].notna()].copy()
    df["_strata"] = df["시점관계"]                       # Class 없으면 시점관계로 층화
    return df


def sample(df: pd.DataFrame, n: int) -> pd.DataFrame:
    """층(strata)별 최소 보장 + 나머지 무작위. 희소 Class도 검수에 포함되게."""
    groups = list(df.groupby("_strata"))
    per = max(1, n // (len(groups) * 2))                 # 층별 최소 쿼터
    picks = []
    for _, g in groups:
        picks.append(g.sample(min(len(g), per), random_state=SEED))
    base = pd.concat(picks)
    rest = df.drop(base.index)
    short = n - len(base)
    if short > 0 and len(rest):
        base = pd.concat([base, rest.sample(min(len(rest), short), random_state=SEED)])
    return base.sample(frac=1, random_state=SEED).reset_index(drop=True)  # 순서 섞기


def write_xlsx(rows: pd.DataFrame, path: str):
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "검수"
    headers = ["commentId", "종목", "작성일", "직전_가격흐름", "댓글",
               "사람_시점관계", "사람_예측_방향", "메모"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="DDDDDD")
    for _, r in rows.iterrows():
        ws.append([r["commentId"], r["종목명"], str(r["작성일"])[:16],
                   context_text(r["종목명"], r["작성일"]), str(r["text"]), "", "", ""])

    # 드롭다운(데이터 검증)
    n = len(rows) + 1
    dv1 = DataValidation(type="list", formula1='"예측,리액션"', allow_blank=True)
    dv2 = DataValidation(type="list", formula1='"상승,하락,없음"', allow_blank=True)
    ws.add_data_validation(dv1); ws.add_data_validation(dv2)
    dv1.add(f"F2:F{n}"); dv2.add(f"G2:G{n}")

    widths = {"A": 14, "B": 7, "C": 17, "D": 46, "E": 50, "F": 13, "G": 13, "H": 18}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    for row in ws.iter_rows(min_row=2, max_row=n):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"

    # 안내 시트
    info = wb.create_sheet("작성법")
    for ln in [
        "■ κ 검수 작성법",
        "1) '댓글'과 '직전_가격흐름'만 보고 판단하세요. (작성일 이후 실제 등락은 모른다고 가정 — 사후판단 금지)",
        "2) [사람_시점관계]: 예측(앞으로 오른다/내린다 단언) / 리액션(이미 움직인 것에 대한 반응).",
        "3) [사람_예측_방향]: 상승 / 하락 / 없음(방향 예측 아님).",
        "4) 색상관례: 빨강·빨간불=상승, 파랑·파란불=하락 (한국식).",
        "5) 애매하면 [메모]에 한 줄. 모르겠으면 비워두기(그 행은 κ 계산서 제외).",
        "6) 다 채우면 저장 후: python3 labeling/compute_kappa.py",
    ]:
        info.append([ln])
    info.column_dimensions["A"].width = 110
    wb.save(path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("-n", type=int, default=N_DEFAULT)
    a = ap.parse_args()
    df = load_pool()
    rows = sample(df, a.n)
    out_xlsx = os.path.join(DATA, "review_sheet.xlsx")
    write_xlsx(rows, out_xlsx)
    # 정답키(숨김): commentId + 파이프라인 라벨 + 텍스트(불일치 분석용)
    keycols = ["commentId", "시점관계", "예측_방향", "text"]
    if "Class" in rows:
        keycols.append("Class")
    rows[keycols].rename(columns={"시점관계": "key_시점관계",
                                  "예측_방향": "key_예측_방향"}).to_csv(
        os.path.join(DATA, "review_key.csv"), index=False, encoding="utf-8-sig")
    print(f"생성: data/review_sheet.xlsx ({len(rows)}건) + data/review_key.csv(정답 숨김)")
    print("→ 시트의 F·G열(사람_시점관계/사람_예측_방향) 채운 뒤: python3 labeling/compute_kappa.py")


if __name__ == "__main__":
    main()
